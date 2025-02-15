import pandas as pd
import openpyxl
from io import BytesIO
from flask import Flask, request, send_file, render_template
import re

app = Flask(__name__, template_folder='templates')

# (Assume you already have your CSV transformation logic that produces df_transposed)
# For this example, let's assume df_transposed is your final DataFrame.

def safe_sum_by_template(df, prefix, suffix, year):
    pattern = re.escape(prefix) + r"EF" + str(year) + r"(?:_[A-Za-z0-9]+)?" + re.escape(suffix)
    total = 0
    for col in df.columns:
        if re.fullmatch(pattern, col):
            total += df[col].sum()
    return total

def transform_csv_data(file_data, start_year, end_year):
    df = pd.read_csv(file_data)
    output_data = {
        "Year": [],
        "Undergrad enrollment - Full Time": [],
        "Undergrad enrollment - Part Time": [],
        "Graduate Enrollment": [],
        "First-Time Freshman Enrollment": [],
        "New Transfer Enrollment (Fall)": [],
        "Total Enrollment (undergrad + grad)": []
    }
    for year in range(start_year, end_year + 1):
        undergrad_full = safe_sum_by_template(df, "Full time total (", "  All students  Undergraduate total)", year)
        undergrad_part = safe_sum_by_template(df, "Part time total (", "  All students  Undergraduate total)", year)
        grad_full = safe_sum_by_template(df, "Full time total (", "  All students  Graduate and First professional)", year)
        grad_part = safe_sum_by_template(df, "Part time total (", "  All students  Graduate and First professional)", year)
        fresh_full = safe_sum_by_template(df, "Full time total (", "  All students  Undergraduate  Degree/certificate-seeking  First-time)", year)
        fresh_part = safe_sum_by_template(df, "Part time total (", "  All students  Undergraduate  Degree/certificate-seeking  First-time)", year)
        transf_full = safe_sum_by_template(df, "Full time total (", "  All students  Undergraduate  Other degree/certificate-seeking  Transfer-ins)", year)
        transf_part = safe_sum_by_template(df, "Part time total (", "  All students  Undergraduate  Other degree/certificate-seeking  Transfer-ins)", year)

        grad_total = grad_full + grad_part
        freshman_total = fresh_full + fresh_part
        transfer_total = transf_full + transf_part
        total_enrollment = undergrad_full + undergrad_part + grad_total

        output_data["Year"].append(year)
        output_data["Undergrad enrollment - Full Time"].append(undergrad_full)
        output_data["Undergrad enrollment - Part Time"].append(undergrad_part)
        output_data["Graduate Enrollment"].append(grad_total)
        output_data["First-Time Freshman Enrollment"].append(freshman_total)
        output_data["New Transfer Enrollment (Fall)"].append(transfer_total)
        output_data["Total Enrollment (undergrad + grad)"].append(total_enrollment)

    df_out = pd.DataFrame(output_data)
    df_out = df_out.set_index("Year").sort_index(ascending=False)
    df_transposed = df_out.transpose()
    desired_order = [
        "Total Enrollment (undergrad + grad)",
        "Undergrad enrollment - Full Time",
        "Undergrad enrollment - Part Time",
        "Graduate Enrollment",
        "First-Time Freshman Enrollment",
        "New Transfer Enrollment (Fall)"
    ]
    df_transposed = df_transposed.loc[desired_order]
    return df_transposed

def write_to_template_in_memory(df, template_path, start_row=6, start_col=4):
    """
    Writes the DataFrame into the Excel template and returns an in-memory BytesIO object.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Write headers
    for j, header in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=header)

    # Write data
    for i, row in enumerate(df.itertuples(index=False), start=start_row+1):
        for j, value in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=value)

    # Save workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files["file"]
        start_year = int(request.form.get("start_year"))
        end_year = int(request.form.get("end_year"))

        # Transform CSV data into a DataFrame
        df_transposed = transform_csv_data(file, start_year, end_year)

        # Write DataFrame to Excel template in memory
        template_path = "my_template.xlsx"  # Make sure this file is in your repo
        excel_io = write_to_template_in_memory(df_transposed, template_path)

        # Return the Excel file as a download
        return send_file(
            excel_io,
            as_attachment=True,
            download_name="transformed_report.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
